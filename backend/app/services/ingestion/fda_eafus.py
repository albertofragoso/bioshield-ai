"""FDA EAFUS ingestion: parse the FDA 'Everything Added to Food' Excel workbook.

Live URL varies over time; we try a list of candidate URLs and fall back to a
bundled fixture (.xlsx) if all live fetches fail. The parser expects columns:
    'CAS Reg No. (or other ID)', 'Substance', 'Document Number', etc.
"""

from __future__ import annotations

import io
import logging
from collections.abc import Iterable
from pathlib import Path

import httpx
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import Settings
from app.models import IngestionLog
from app.services.ingestion.common import (
    IngestionRecord,
    checksum,
    finish_log,
    get_or_create_source,
    index_record,
    start_log,
    upsert_ingredient,
    upsert_regulatory_status,
)

logger = logging.getLogger(__name__)

SOURCE_NAME = "FDA_EAFUS"
FIXTURE_PATH = Path(__file__).parents[3] / "data" / "seed" / "fda_eafus_fixture.xlsx"

# FDA rotates this URL; try candidates in order.
CANDIDATE_URLS = [
    "https://www.cfsanappsexternal.fda.gov/scripts/fdcc/eafus.xlsx",
    "https://www.cfsanappsexternal.fda.gov/scripts/fdcc/index.cfm?set=EAFUS&exportopt=default&Qformat=EXCEL_SLIMVALUE",
    "https://www.cfsanappsexternal.fda.gov/scripts/fdcc/?set=EAFUS&exportopt=default&Qformat=EXCEL_SLIMVALUE",
]


class FDAFetchError(RuntimeError):
    """Raised when all candidate URLs fail."""


def parse_workbook(raw_bytes: bytes) -> list[IngestionRecord]:
    workbook = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
    sheet = workbook.active

    headers: list[str] = []
    records: list[IngestionRecord] = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h or "").strip() for h in row]
            continue
        if not row:
            continue
        row_dict = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        substance = (row_dict.get("Substance") or row_dict.get("Name") or "").strip()
        cas = (row_dict.get("CAS Reg No. (or other ID)") or row_dict.get("CAS") or "").strip()
        if not substance:
            continue
        records.append(
            IngestionRecord(
                canonical_name=substance,
                cas_number=cas if cas and "-" in cas else None,
                status="APPROVED",
                hazard_note=None,
            )
        )
    return records


async def fetch_live_bytes(client: httpx.AsyncClient) -> bytes:
    """Try candidate URLs in order; raise FDAFetchError if all fail."""
    last_exc: Exception | None = None
    for url in CANDIDATE_URLS:
        try:
            response = await client.get(url)
            response.raise_for_status()
            content = response.content
            # Sanity check: xlsx files start with PK (zip magic bytes)
            if content[:2] == b"PK":
                logger.info("FDA EAFUS fetched from %s (%d bytes)", url, len(content))
                return content
            logger.warning("FDA URL %s returned non-xlsx content, skipping", url)
        except httpx.HTTPStatusError as exc:
            logger.warning("FDA URL %s → %s", url, exc.response.status_code)
            last_exc = exc
        except httpx.RequestError as exc:
            logger.warning("FDA URL %s network error: %s", url, exc)
            last_exc = exc
    raise FDAFetchError("All FDA EAFUS candidate URLs failed") from last_exc


async def run(
    db: Session,
    settings: Settings,
    records: Iterable[IngestionRecord] | None = None,
    raw_bytes: bytes | None = None,
) -> IngestionLog:
    """Ingest FDA EAFUS records into SQL + Chroma.

    Priority: records > raw_bytes > live fetch > bundled fixture.
    """
    if records is None:
        if raw_bytes is None:
            try:
                async with httpx.AsyncClient(timeout=60) as client:
                    raw_bytes = await fetch_live_bytes(client)
            except FDAFetchError:
                logger.warning("FDA live fetch failed; using bundled fixture at %s", FIXTURE_PATH)
                raw_bytes = FIXTURE_PATH.read_bytes()
        records = parse_workbook(raw_bytes)

    records_list = list(records)
    source_bytes = raw_bytes or str(len(records_list)).encode()
    source = get_or_create_source(
        db,
        name=SOURCE_NAME,
        region="US",
        version="EAFUS_latest",
        source_checksum=checksum(source_bytes),
        license_="Public Domain",
        format_="XLSX",
    )
    log = start_log(db, source, checksum(source_bytes))

    for rec in records_list:
        ing = upsert_ingredient(db, rec)
        upsert_regulatory_status(db, ing, source, rec)
        await index_record(ing, rec, SOURCE_NAME, settings)

    finish_log(log, records_processed=len(records_list), status="SUCCESS")
    db.commit()
    logger.info("FDA EAFUS ingested: %d records", len(records_list))
    return log
