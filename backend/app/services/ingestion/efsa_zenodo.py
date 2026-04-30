"""EFSA OpenFoodTox ingestion via Zenodo API.

Zenodo record 8120114 contains Excel files. We download
SubstanceCharacterisation_KJ_2023.xlsx to extract substance names, CAS, and
EC reference numbers. Status defaults to UNDER_REVIEW (EFSA assessments are
ongoing risk evaluations, not binary approved/banned decisions).
"""

from __future__ import annotations

import io
import logging
import re
from collections.abc import Iterable

import httpx
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import Settings
from app.models import IngestionLog
from app.services.ingestion.common import (
    IngestionRecord,
    checksum,
    finish_log,
    get_or_create_source,
    index_record,
    start_log,
    upsert_ingredient,
    upsert_regulatory_status,
)

logger = logging.getLogger(__name__)

SOURCE_NAME = "EFSA_OpenFoodTox"
ZENODO_RECORD_URL = "https://zenodo.org/api/records/8120114"
TARGET_FILE = "SubstanceCharacterisation_KJ_2023.xlsx"

_XML_ENTITY_RE = re.compile(r"_x([0-9A-Fa-f]{4})_")


def _decode_name(raw: str) -> str:
    """Decode Excel XML-escaped characters like _x0028_ → ("""
    return _XML_ENTITY_RE.sub(lambda m: chr(int(m.group(1), 16)), raw).strip()


def parse_workbook_bytes(raw_bytes: bytes) -> list[IngestionRecord]:
    wb = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active

    records: list[IngestionRecord] = []
    seen: set[str] = set()

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header row
        if not row or not row[0]:
            continue

        relation = str(row[1] or "").strip()
        if relation != "as_x0020_such" and relation != "as such":
            continue  # skip metabolite / component rows

        name = _decode_name(str(row[0]))
        if not name or name in seen:
            continue
        seen.add(name)

        cas = _decode_name(str(row[3] or "")).strip() or None
        ec_ref = _decode_name(str(row[4] or "")).strip() or None

        if cas and "-" not in cas:
            cas = None

        records.append(
            IngestionRecord(
                canonical_name=name,
                cas_number=cas,
                e_number=ec_ref,
                status="UNDER_REVIEW",
                hazard_note=None,
                evaluated_at=None,
            )
        )

    return records


async def fetch_live_bytes(client: httpx.AsyncClient) -> bytes:
    meta = (await client.get(ZENODO_RECORD_URL)).json()
    files = meta.get("files", [])
    file_url = next((f["links"]["self"] for f in files if f["key"] == TARGET_FILE), None)
    if not file_url:
        raise ValueError(f"{TARGET_FILE} not found in Zenodo record {ZENODO_RECORD_URL}")
    response = await client.get(file_url)
    response.raise_for_status()
    return response.content


async def run(
    db: Session,
    settings: Settings,
    records: Iterable[IngestionRecord] | None = None,
    payload: dict | None = None,
) -> IngestionLog:
    raw_bytes: bytes | None = None

    if records is None:
        async with httpx.AsyncClient(timeout=120) as client:
            raw_bytes = await fetch_live_bytes(client)
        records = parse_workbook_bytes(raw_bytes)

    records_list = list(records)
    src_bytes = raw_bytes or str(len(records_list)).encode()

    source = get_or_create_source(
        db,
        name=SOURCE_NAME,
        region="EU",
        version="OpenFoodTox_KJ_2023",
        source_checksum=checksum(src_bytes),
        license_="CC BY 4.0",
        format_="XLSX",
    )
    log = start_log(db, source, checksum(src_bytes))

    for rec in records_list:
        ing = upsert_ingredient(db, rec)
        upsert_regulatory_status(db, ing, source, rec)
        await index_record(ing, rec, SOURCE_NAME, settings)

    finish_log(log, records_processed=len(records_list), status="SUCCESS")
    db.commit()
    logger.info("EFSA OpenFoodTox ingested: %d records", len(records_list))
    return log
